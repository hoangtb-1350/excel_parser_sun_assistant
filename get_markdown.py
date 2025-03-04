# TODO: Hyper, strikethough
# Utils
import re
import math
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from sklearn.metrics import classification_report, confusion_matrix

mode = 'NORMAL'
table_labels = ['HEADER', 'ATTRIBUTE', 'DATA', 'NONE']

class Excel_to_markdown():
    # Predefined color dictionary (add more as needed)
    colors = {
        'Red': (255, 0, 0),
        'Green': (0, 255, 0),
        'Blue': (0, 0, 255),
        'Yellow': (255, 255, 0),
        'Cyan': (0, 255, 255),
        'Magenta': (255, 0, 255),
        'Black': (0, 0, 0),
        'White': (255, 255, 255),
        'Gray': (128, 128, 128),
        'Orange': (255, 165, 0),
        'Purple': (128, 0, 128),
        'Pink': (255, 192, 203),
        'Brown': (165, 42, 42)
    }

    # Function to calculate Euclidean distance between two RGB colors
    @staticmethod
    def color_distance(c1, c2):
        return math.sqrt(sum((e1 - e2) ** 2 for e1, e2 in zip(c1, c2)))
    
    @staticmethod
    def get_table_as_array(md_content):
        lines = md_content.strip().split("\n")
        table = []
        
        for line in lines:
            if line.startswith("|---"):  # Skip the separator line
                continue
            
            row = [cell.strip() for cell in line.split("|") if cell.strip()]
            table.append(row)
        return table

    @staticmethod
    def get_trim_markdown_table(md_content, output_range):
        """ 
        Get the trimmed markdown table
        md_content: Markdown table content
        Output range: [start_row_idx, end_row_idx, start_col_idx, end_col_idx]
        """
        table = Excel_to_markdown.get_table_as_array(md_content)
        row_range = output_range[:2]
        col_range = output_range[2:]
        for r_idx, r in enumerate(table):
            if r_idx < row_range[0] or r_idx > row_range[1]:
                continue
            for c_idx, cell in enumerate(r):
                if c_idx < col_range[0] or c_idx > col_range[1]:
                    continue
                md_content += f"|{cell}"
            md_content += "|\n"
        return md_content
        
    @staticmethod
    def evaluate_markdown_table(md_pred, md_label):
        """ 
        Evaluate the predicted markdown table
        """
        pred_table = Excel_to_markdown.get_table_as_array(md_pred)
        label_table = Excel_to_markdown.get_table_as_array(md_label)
        num_rows = len(pred_table)
        num_cols = len(pred_table[0])
        
        if num_rows != len(label_table) or num_cols != len(label_table[0]):
            raise ValueError("The number of rows or columns in the predicted table does not match the label table.")
        
        preds = [j for i in pred_table for j in i]
        labels = [j for i in label_table for j in i]
        report = classification_report(labels, preds, output_dict=True, labels=table_labels)
        confusion = confusion_matrix(labels, preds, labels=table_labels)
        return report, confusion


    # Find the nearest color
    def find_nearest_color(target_rgb):
        target_rgb = tuple(int(target_rgb[i:i+2], 16) for i in (0, 2, 4))
        nearest_color = None
        min_distance = float('inf')

        for name, rgb in Excel_to_markdown.colors.items():
            distance = Excel_to_markdown.color_distance(target_rgb, rgb)
            if distance < min_distance:
                min_distance = distance
                nearest_color = name

        return nearest_color

    def get_dropdown_values(self, cell):
        for dv in self.ws.data_validations.dataValidation:
            # Check if the cell is in the data validation range
            try:
                check_cell_in = cell in dv.cells
            except Exception:
                return False
            is_dropdown = dv.type == 'list'
            if not is_dropdown or not check_cell_in:
                continue
            if cell in dv.cells:
                # Check if it's a list-based dropdown
                if dv.type == "list":
                    # If it's a direct list (e.g., "Option1,Option2")
                    if dv.formula1.startswith('"'):
                        return dv.formula1.strip('"').split(',')
                    else:
                        # If it's a range reference (e.g., =Sheet1!$D$1:$D$5)
                        try:
                            ref = dv.formula1.replace('=', '')
                            values = [c[0].value for c in self.ws[ref]]
                            if len(values) > 3:
                                values = values[:3] + ["..."]
                            return [v for v in values if v is not None]
                        except: # Other type, not dropdown
                            return False
        return False

    def get_value(self, cell, merge_cell_values, image_cell_values, get_comment=False, get_color=False):
        def _get_hyperlink(text):
            pattern = r'=HYPERLINK\("([^"]+)"\s*,\s*"([^"]+)"\)'
            match_result = re.search(pattern, text)

            if match_result:
                url = match_result.group(1)
                display_text = match_result.group(2)
                return f"[{display_text}]({url})"
            else:
                return text

        # Check pure value
        value = ""
        cell_instance = None
        if cell.value is not None and type(cell.value) == str:
            # Check if it's label
            if cell.value == "HEADER" or cell.value == 'ATTRIBUTE' and self.get_label:
                return cell.value
            # Check hyperlink in case of formula
            no_compile_cell = self.no_compile_ws[cell.coordinate]
            if no_compile_cell.value[:10] == '=HYPERLINK':
                value = _get_hyperlink(no_compile_cell.value)
            else:
                value = str(cell.value)
            cell_instance = cell
        elif cell.value is not None and type(cell.value) in [int, float]:
            value = str(cell.value)
            cell_instance = cell
        elif cell.coordinate in merge_cell_values and merge_cell_values[cell.coordinate].value is not None and type(merge_cell_values[cell.coordinate].value) == str and self.merge_strategy == 'duplicate':
            no_compile_cell = self.no_compile_ws[merge_cell_values[cell.coordinate].coordinate]
            if no_compile_cell.value[:10] == '=HYPERLINK':
                value = _get_hyperlink(no_compile_cell.value)
            else:
                value = str(merge_cell_values[cell.coordinate].value)
            cell_instance = merge_cell_values[cell.coordinate]

            if value == "HEADER" or value == 'ATTRIBUTE' and self.get_label:
                return value # Return immediately

        # check if dropdown
        if cell_instance is not None and self.get_dropdown_values(cell_instance.coordinate):

            value += " from a list: " + ", ".join(self.get_dropdown_values( cell_instance.coordinate))

        # check comment
        if get_comment and cell.comment:
            comment_break = cell.comment.text.split(":")
            if len(comment_break) == 1:
                value += f". Comment: {cell.comment.text}"
            else:
                value += f". Comment: {' '.join(comment_break[1:])}"

        # check in case of image
        if cell.coordinate in image_cell_values and image_cell_values[cell.coordinate] is not None:
            value += " " + str(image_cell_values[cell.coordinate].value)
            # No cell instance as we get from img

        # check property
        if cell_instance is not None:
            # Check property
            if getattr(cell_instance, 'font', False):
                if cell_instance.font.strike:
                    value = f"~~{value}~~"
                if cell_instance.font.bold:
                    value = f"**{value}**"
                if cell_instance.font.italic:
                    value = f"*{value}*"

            if getattr(cell_instance, 'hyperlink', False):
                value = f"[{value}]({str(cell_instance.hyperlink.target)})"

            # Ưu tiên color nền hơn
            if get_color:
                set_color = False
                if getattr(cell_instance.fill, 'start_color', False):
                    if cell_instance.fill.start_color.rgb and 'Values must be of' not in str(cell_instance.fill.start_color.rgb) and cell_instance.fill.start_color.rgb[2:] != '000000':
                        value = f"<span style=\"color:{Excel_to_markdown.find_nearest_color(cell_instance.fill.start_color.rgb[2:])}\">{value}</span>"
                        set_color= True
                if getattr(cell_instance, 'font', False) and getattr(cell_instance.font, 'color', False) and getattr(cell_instance.font.color, 'rgb', False) and not set_color:
                    try:
                        if cell_instance.font.color.rgb and 'Values must be of' not in str(cell_instance.font.color.rgb) and cell_instance.font.color.rgb[2:] != '000000':
                            value = f"<span style=\"color:{Excel_to_markdown.find_nearest_color(cell_instance.font.color.rgb[2:])}\">{value}</span>"
                    except Exception as e:
                        print(e)
        if cell.coordinate == 'L166':
            print('value:', value)
        if self.get_label:
            if value == "":
                return "NONE"
            else:
                return "DATA"

        if value == "":
            return " "
        return value.replace("\n", "<br>")

    def check_in_range(self, range_a, range_b):
        """
        True if range_a completely in range_b
        """
        min_col_a, min_row_a, max_col_a, max_row_a = range_boundaries(range_a)
        min_col_b, min_row_b, max_col_b, max_row_b = range_boundaries(range_b)

        # Check if range_a is within range_b
        is_within = (min_col_b <= min_col_a <= max_col_b and
                    min_col_b <= max_col_a <= max_col_b and
                    min_row_b <= min_row_a <= max_row_b and
                    min_row_b <= max_row_a <= max_row_b)
        return is_within

    def convert_table_to_markdown(self, file_name, sheet, range, out_type, merge_strategy='duplicate', get_label=False, output_range=[], max_sentences = False, max_token = False):
        """        
        Converts a specified range of an Excel sheet to markdown or HTML format.

        Parameters:
        file_name (str): The name of the Excel file.
        sheet (str or int): The sheet name or index.
        range (str): The range of the table to convert, e.g., 'A1:C5'.
        out_type (str): The output type, either 'html' or 'markdown'.
        merge_strategy (str, optional): Strategy for handling merged cells, either 'duplicate' or 'ignore'. Default is 'duplicate'.
        get_label (bool, optional): Whether to get labels. Default is False.
        output_range (list, optional): Specific range to export markdown, e.g., [1,3,2,5] as start_row_idx, end_row_idx, start_col_idx, end_col_idx (inclusive). Default is an empty list.
        max_sentences (int | bool, optional): Maximum number of sentences. Default is False as no restriction.
        max_token (int | bool, optional): Maximum number of tokens. Default is False as no restriction.

        Returns:
        str: The converted table in markdown or HTML format.
        """
        self.merge_strategy = merge_strategy
        self.get_label = get_label
        # Find rows, cols number
        self.wb = load_workbook(file_name, data_only=True)
        self.no_compile_wb = load_workbook(file_name, data_only=False)
        if type(sheet) == int:
            self.ws = self.wb.worksheets[sheet]
            self.no_compile_ws = self.no_compile_wb.worksheets[sheet]
        else:
            self.ws = self.wb[sheet]
            self.no_compile_ws = self.no_compile_wb[sheet]

        cell_range = self.ws[range]
        no_compile_cell_range = self.no_compile_ws[range]

        num_rows = len(cell_range)
        num_cells = sum(len(row) for row in cell_range)

        # get merge cell list
        list_merge_cells = [merged_range.coord for merged_range in self.ws.merged_cells.ranges if self.check_in_range(merged_range.coord, range)]
        merge_cell_values = {}
        for merge_cell in list_merge_cells:
            merge_cell_range = self.ws[merge_cell]
            # breakpoint()
            for r in merge_cell_range:
                for cell in r:
                    if cell.value is not None:
                        merge_cell_values[cell.coordinate] = cell
                    else:
                        merge_cell_values[cell.coordinate] = merge_cell_range[0][0]

        if mode=='DEBUG':
            print(merge_cell_values)

        # Rm duplicate row/col
        duplicate_row_idx = []
        duplicate_col_idx = []
        row_set = set()
        col_set = set()
        for r_idx, r in enumerate(cell_range):
            bf_len = len(row_set)
            row_value = " ".join(str(c.value) for c in r)
            if re.fullmatch(r'(?:None\s*)*', row_value):
                duplicate_row_idx.append(r_idx)
                continue
            row_set.add(row_value)
            # if len(row_set) == bf_len:
            #     duplicate_row_idx.append(r_idx)

        for c_idx, c in enumerate(zip(*cell_range)):
            bf_len = len(col_set)
            col_value = " ".join(str(c.value) for c in c)
            if re.fullmatch(r'(?:None\s*)*', col_value):
                duplicate_col_idx.append(c_idx)
                continue
            col_set.add(col_value)
            # if len(col_set) == bf_len:
            #     duplicate_col_idx.append(c_idx)

        if mode=="DEBUG":
            print(duplicate_row_idx)
            print(duplicate_col_idx)

        # Define output range
        if len(output_range) == 0:
            row_range = 0, 999999
            col_range = 0, 999999
        else:
            row_range = [output_range[0], output_range[1]]
            col_range = [output_range[2], output_range[3]]
        # Handle as markdown
        if out_type=='markdown':
            md_content = ""
            header = True
            r_count = 0
            for r_idx, r in enumerate(cell_range):
                if r_idx in duplicate_row_idx:
                    continue
                if r_count < row_range[0] or r_count > row_range[1]:
                    r_count += 1
                    continue
                r_count += 1
                c_count = 0
                for c_idx, cell in enumerate(r):
                    if c_idx in duplicate_col_idx:
                        continue
                    if c_count < col_range[0] or c_count > col_range[1]:
                        c_count += 1
                        continue
                    c_count += 1
                    # print(cell.coordinate)
                    md_content += f"|{self.get_value(cell, merge_cell_values, {})}"
                md_content += "|\n"
            return md_content

def flatten_dict(nested_dict):
    result = {}

    def helper(current_dict, key_prefix):
        for key, value in current_dict.items():
            new_key = f"{key_prefix}{key}" if key_prefix else key
            if isinstance(value, dict):
                helper(value, new_key + ".")
            else:
                result[new_key] = value

    helper(nested_dict, "")
    return result


